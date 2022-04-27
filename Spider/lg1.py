from selenium import webdriver
import time
from openpyxl import load_workbook, Workbook
from selenium.webdriver.chrome.options import Options
import os

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--disable-dev-shm-usage')
chromedriver = "/usr/bin/chromedriver"
os.environ["webdriver.chrome.driver"] = chromedriver

file = './Spider/a.xlsx'
max_a = load_workbook(file).active.max_row
print(max_a)


def writeexcle(t):
    wb = load_workbook(file)
    sheet = wb.active
    max_row = sheet.max_row - t

    row_max = 'a' + str(max_row)
    print(sheet)

    if max_row > 0:
        a = sheet[row_max].value
        print(a)
        return a
    else:
        print("空了")
        return None


# options = webdriver.FirefoxOptions()
# options.set_headless(True)
# options.add_argument("--headless")  # 设置火狐为headless无界面模式
# options.add_argument("--disable-gpu")
# d = webdriver.Firefox(options=options)

# d = webdriver.Firefox()
# # d = webdriver.Chrome()
# d.implicitly_wait(5)


def rw():
    time.sleep(1)


def dl(phone, d):
    d.get('https://www.chaojijishi.com/h5/#/pages/login/login?from=user')
    phh = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[1]/uni-input/div/input')

    rw()
    phh.send_keys(phone)

    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[2]/uni-input/div/input').send_keys(
        123456)
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[3]/uni-view[1]/uni-view').click()
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[5]/uni-view/uni-view').click()
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[5]/uni-view[4]').click()
    rw()


def zz(d, mm):
    rw()
    d.get('https://www.chaojijishi.com/h5/#/pages/subpack1/pay-model/JDpay')
    rw()

    jifen = d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view/uni-swiper/div/div/div/uni-swiper-item/uni-view/uni-view/uni-view[2]').text

    print(jifen)
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view/uni-view[1]/uni-view[1]').click()
    rw()
    # 输入号码
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[2]/uni-view/uni-view/uni-view[2]/uni-view[2]/uni-input/div/input').send_keys(
        17024465574)
    rw()
    # 输入积分
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[2]/uni-view/uni-view/uni-view[3]/uni-view[2]/uni-input/div/input').send_keys(
        jifen)

    # 输入密码
    rw()
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[2]/uni-view/uni-view/uni-view[4]/uni-view[2]/uni-input/div/input').send_keys(
        mm)
    # 确定
    rw()
    # input('测试')
    d.find_element_by_xpath(
        '/html/body/uni-app/uni-page/uni-page-wrapper/uni-page-body/uni-view/uni-view[2]/uni-view[2]/uni-view/uni-view/uni-view[5]/uni-view[2]').click()
    rw()
    return jifen


def cs():
    a = max_a
    c = 0
    zong = 0
    while a > 0:
        d = webdriver.Chrome(chrome_options=chrome_options, executable_path=chromedriver)
        # d = webdriver.Chrome()
        # d = webdriver.Firefox()
        d.implicitly_wait(5)
        dl(writeexcle(c), d)
        mm = "123456"
        zong += float(zz(d, mm))
        a -= 1
        d.close()

    print("总共")
    print(zong)


cs()
